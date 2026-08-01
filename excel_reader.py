from openpyxl import load_workbook

def get_day_menu(day):
    workbook = load_workbook("menu.xlsx")
    sheet = workbook.active

    menu = {}

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[0] == day:
            menu = {
                "Breakfast": row[1],
                "Lunch": row[2],
                "Snacks": row[3],
                "Dinner": row[4]
            }
            break

    workbook.close()
    return menu