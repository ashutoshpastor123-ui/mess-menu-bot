from apscheduler.schedulers.background import BackgroundScheduler
from telegram import Bot
from telegram.error import TelegramError
from dotenv import load_dotenv
from openpyxl import load_workbook

import os
import asyncio
import logging
from datetime import datetime

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.getenv("BOT_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")

if not BOT_TOKEN or not CHAT_ID:
    raise ValueError("BOT_TOKEN and CHAT_ID must be set")

bot = Bot(BOT_TOKEN)

scheduler = BackgroundScheduler()

DAY_COLUMNS = {
    "Monday": 3,
    "Tuesday": 4,
    "Wednesday": 5,
    "Thursday": 6,
    "Friday": 7,
    "Saturday": 8,
    "Sunday": 9
}

MENU_DATA = {}


def load_menu():
    global MENU_DATA

    workbook = load_workbook("menu.xlsx")
    sheet = workbook.active

    MENU_DATA = {}


    for row in sheet.iter_rows(min_row=3, values_only=True):

        day = row[0]

        if not day:
            continue

        MENU_DATA[day] = {
            "BREAKFAST": row[1],
            "LUNCH": row[2],
            "SNACKS": row[3],
            "DINNER": row[4]
        }

    workbook.close()

    logger.info("✅ Menu Loaded")


async def send_reminder(meal):

    try:

        today = datetime.now().strftime("%A")

        menu = MENU_DATA[today].get(meal)

        if not menu:
            return

        await bot.send_message(
            chat_id=CHAT_ID,
            text=f"""🔔 {meal.title()} starts in 10 minutes!

🍽 Today's Menu

{menu}
"""
        )

        logger.info(f"{meal} reminder sent")

    except TelegramError as e:
        logger.error(e)


def breakfast():
    asyncio.run(send_reminder("BREAKFAST"))


def lunch():
    asyncio.run(send_reminder("LUNCH"))


def snacks():
    asyncio.run(send_reminder("SNACKS"))


def dinner():
    asyncio.run(send_reminder("DINNER"))


def start_scheduler():

    load_menu()

    scheduler.add_job(load_menu, "cron", hour=0, minute=0)

    scheduler.add_job(breakfast, "cron", hour=7, minute=50)

    scheduler.add_job(lunch, "cron", hour=12, minute=20)

    scheduler.add_job(snacks, "cron", hour=16, minute=20)

    scheduler.add_job(dinner, "cron", hour=20, minute=5)

    scheduler.start()

    logger.info("🤖 Scheduler Started")


def stop_scheduler():

    scheduler.shutdown()

    logger.info("🛑 Scheduler Stopped")